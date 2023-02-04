import os
import openpyxl
from datetime import date, datetime
from errorHandler import *


def getPath():

    # Give the location of the file
    osPath = os.getcwd()  # get path of this directory
    file = '\delivery.xlsx'  # filename
    return osPath+file


def openExcel():

    try:
        path = getPath()

        # To open the workbook
        # workbook object is created
        workbook = openpyxl.load_workbook(path)
        return workbook

    except Exception as e:

        title = 'De code is vastgelopen in de openExcel functie'
        error = e
        SendErrorMail(title, error)
        input()


def readDataFromExcel(workbook, data):

    try:
        # Get sheet names
        sheet = workbook.active

        # Read a Cell
        # Get cell value
        cell_obj = sheet.cell(row=1, column=1)

        # Print value of cell
        print(cell_obj.value)

    except Exception as e:

        title = 'De code is vastgelopen in de readDataFromExcel functie'
        error = e
        SendErrorMail(title, error)
        input()


# This function delete all the rows in excel with the same ordernummer and artikelcode that was found in de emails
def deleteArticleDubbel(workbook, articleObjectList):

    try:
        # Get sheet names
        sheet = workbook.active
        index = 1
        del_rows = []

        for row in sheet.iter_rows(min_row=2):
            index += 1
            orderNumber = row[0].value
            articleId = row[1].value
            number = row[2].value

            # check dubbel with the inportinglist
            for articleObject in articleObjectList:
                if orderNumber == articleObject.orderNumber and articleId == articleObject.articleId and number == articleObject.number:
                    # row matches object
                    del_rows.append(index)
                    break

        for r in reversed(del_rows):
            sheet.delete_rows(r)
        index = 1

        path = getPath()
        # Save the file
        workbook.save(path)

    except Exception as e:

        title = 'De code is vastgelopen in de deleteArticleDubbel functie'
        error = e
        SendErrorMail(title, error)
        input()


def checkOnDubbel(workbook):
    try:
        sheet = workbook.active
        index = 1
        del_rows = []

        for row in sheet.iter_rows(min_row=2):
            index += 1
            orderNumber = row[0].value
            articleId = row[1].value
            number = row[2].value

            # check dubbel in excel it self
            for row in sheet.iter_rows(min_row=index+1):

                orderNumberCheck = row[0].value
                articleIdCheck = row[1].value
                numberCheck = row[2].value

                if orderNumber == orderNumberCheck and articleId == articleIdCheck and number == numberCheck:
                    # row matches object
                    del_rows.append(index)
                    break

        for r in reversed(del_rows):
            sheet.delete_rows(r)
        index = 1

        path = getPath()
        # Save the file
        workbook.save(path)

    except Exception as e:

        title = 'De code is vastgelopen in de checkOnDubbel functie'
        error = e
        SendErrorMail(title, error)
        input()


def deleteArticlesRowExpireddate(workbook):
    try:
        today = date.today()

        dateTimeNow = datetime.now()

        # Get sheet names
        sheet = workbook.active

        index = 1
        del_rows = []

        # there is an issue with the first row. because is a tekst and not a Date
        # So this loop has to start from the second row.
        for row in sheet.iter_rows(min_row=2):
            index += 1
            deliverDateStr = row[3].value

            if not deliverDateStr == "-":
                deliverDate = datetime.strptime(deliverDateStr, "%d/%m/%y")

                if dateTimeNow > deliverDate:
                    # If deliverDate is expired with current day
                    del_rows.append(index)

        for r in reversed(del_rows):
            sheet.delete_rows(r)

        path = getPath()
        # Save the file
        workbook.save(path)

    except Exception as e:

        title = 'De code is vastgelopen in de deleteArticlesRowExpireddate functie'
        error = e
        SendErrorMail(title, error)
        input()


def writeDataToExcel(workbook, articleObjectList):
    try:
        # Get sheet names
        sheet = workbook.active

        for articleObject in articleObjectList:
            max_row = sheet.max_row  # find last row of worksheet
            # Write to a Cell
            # write to a particular cell
            sheet.cell(row=max_row+1,
                       column=1).value = articleObject.orderNumber
            sheet.cell(row=max_row+1, column=2).value = articleObject.articleId
            sheet.cell(row=max_row+1, column=3).value = articleObject.number
            sheet.cell(row=max_row+1,
                       column=4).value = articleObject.deliveryDate
            max_row = max_row+1

        path = getPath()
        # Save the file
        workbook.save(path)

    except Exception as e:

        title = 'De code is vastgelopen in de writeDataToExcel functie'
        error = e
        SendErrorMail(title, error)
        input()
